"""
API Reports - Report generation, machine config Excel/HTML export.
"""

import io
import json
import os
import html as _html_mod
from datetime import datetime
from flask import request, jsonify, Response, send_file

from .api_common import check_auth


def register(app, core):
    """Register report routes."""

    @app.route("/api/reports/generate", methods=["POST"])
    def api_generate_report():
        username, err, code = check_auth("settings")
        if err: return err, code
        data = request.json
        report_type = data.get("type", "daily")
        filepath = core.reporting.generate_html_report(report_type=report_type)
        core.db.insert_audit_log(username, "generate_report", f"Type: {report_type}", request.remote_addr)
        return jsonify({"success": True, "path": filepath})

    @app.route("/api/reports/machine-config-export", methods=["POST"])
    def api_machine_config_export():
        username, err, code = check_auth("api")
        if err: return err, code
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return jsonify({"success": False, "error": "openpyxl not installed. Run: pip install openpyxl"}), 500

        machines = core.db.get_machines()
        if not machines:
            return jsonify({"success": False, "error": "No machines found"}), 404

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cau hinh may tram"

        hdr_font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
        hdr_fill = PatternFill(start_color="1A3A5A", end_color="1A3A5A", fill_type="solid")
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_align = Alignment(vertical="top", wrap_text=True)
        swarm_cell_align = Alignment(vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC")
        )

        headers = [
            "Hostname", "Người dùng", "Mã NV", "Email", "TT",
            "Hệ điều hành", "CPU", "RAM\n(GB)", "Ổ đĩa\n(GB)", "GPU",
            "Agent", "IP", "SL\nPM", "📦 Danh sách phần mềm"
        ]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = hdr_align
            c.border = thin_border

        sw_font = Font(name="Consolas", size=9, color="2A4A6A")
        sw_dim_font = Font(name="Consolas", size=9, color="999999", italic=True)

        row = 2
        for m in machines:
            mid = m.get("machine_id", "")
            hw_info = core.db.get_hardware_info(mid)
            user_info = core.db.get_machine_user(mid) or {}
            cfg = hw_info.get("data", {}) if hw_info else {}
            os_info = cfg.get("os", {})
            cpu_info = cfg.get("cpu", {})
            ram_info = cfg.get("ram", {})
            disks = cfg.get("disks", [])
            gpus = cfg.get("gpu", [])
            software_list = cfg.get("installed_software", [])
            total_disk = sum(d.get("size_gb", 0) for d in disks) if disks else 0
            gpu_names = ", ".join(g.get("name", "") for g in (gpus or []) if g.get("name"))

            if software_list:
                sw_lines = []
                for i, sw in enumerate(software_list, 1):
                    name = sw.get("name", "?")
                    ver = sw.get("version", "")
                    pub = sw.get("publisher", "")
                    date = sw.get("install_date", "")
                    ver_str = f" (v{ver})" if ver else ""
                    pub_str = f" - {pub}" if pub else ""
                    date_str = f" [{date}]" if date else ""
                    sw_lines.append(f"{i}. {name}{ver_str}{pub_str}{date_str}")
                sw_text = "\n".join(sw_lines)
                sw_count = len(software_list)
                sw_cell_font = sw_font
            else:
                sw_text = "(Chưa có dữ liệu)"
                sw_count = 0
                sw_cell_font = sw_dim_font

            values = [
                m.get("hostname", mid),
                user_info.get("user_name", ""),
                user_info.get("employee_id", ""),
                user_info.get("email", ""),
                "Online" if m.get("is_online") == 1 else "Offline",
                f"{os_info.get('name', '')} {os_info.get('release', '')} (Build {os_info.get('build', '')})".strip(),
                cpu_info.get("name", ""),
                ram_info.get("total_gb", 0) or 0,
                total_disk if total_disk > 0 else "",
                gpu_names,
                m.get("version", ""),
                m.get("ip_address", ""),
                sw_count,
                sw_text
            ]
            for col, val in enumerate(values, 1):
                c = ws.cell(row=row, column=col, value=val if val is not None else "")
                c.alignment = cell_align
                c.border = thin_border
                c.font = Font(name="Segoe UI", size=10)
            c_sw = ws.cell(row=row, column=14)
            c_sw.font = sw_cell_font
            c_sw.alignment = swarm_cell_align
            if software_list:
                ws.row_dimensions[row].height = max(20, 14 * len(software_list) + 8)
            row += 1

        widths = [18, 18, 8, 25, 6, 30, 26, 7, 8, 22, 10, 14, 5, 52]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row - 1}"
        ws.freeze_panes = "A2"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        core.db.insert_audit_log(username, "export_config_report",
            f"Exported {len(machines)} machines config to Excel", request.remote_addr)

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"GIAM-SAT_Config_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

    @app.route("/api/reports/machine-config-html", methods=["POST"])
    def api_machine_config_html():
        username, err, code = check_auth("api")
        if err: return err, code

        machines = core.db.get_machines()
        if not machines:
            return jsonify({"success": False, "error": "No machines found"}), 404

        online_count = sum(1 for m in machines if m.get("is_online") == 1)
        now = datetime.now().strftime("%d/%m/%Y %H:%M")

        cards_html = ""
        status_colors = {"Online": "#00d4aa", "Offline": "#ff4444"}

        for m in machines:
            mid = m.get("machine_id", "")
            hostname = m.get("hostname", mid)
            hw_info = core.db.get_hardware_info(mid)
            user_info = core.db.get_machine_user(mid) or {}
            cfg = hw_info.get("data", {}) if hw_info else {}
            os_info = cfg.get("os", {})
            cpu_info = cfg.get("cpu", {})
            ram_info = cfg.get("ram", {})
            disks = cfg.get("disks", [])
            gpus = cfg.get("gpu", [])
            software_list = cfg.get("installed_software", [])
            total_disk = sum(d.get("size_gb", 0) for d in disks) if disks else 0
            gpu_names = ", ".join(g.get("name", "") for g in (gpus or []) if g.get("name"))
            status = "Online" if m.get("is_online") == 1 else "Offline"
            sw_count = len(software_list)

            sw_table = ""
            if software_list:
                sw_rows = ""
                for i, sw in enumerate(software_list, 1):
                    name = _html_mod.escape(sw.get("name", "?"))
                    ver = _html_mod.escape(sw.get("version", ""))
                    pub = _html_mod.escape(sw.get("publisher", ""))
                    date = _html_mod.escape(sw.get("install_date", ""))
                    sw_rows += f"<tr><td>{i}</td><td>{name}</td><td>{ver}</td><td>{pub}</td><td>{date}</td></tr>"
                sw_table = f'''<div class="sw-section">
<div class="sw-header">📦 Danh sách phần mềm đã cài đặt ({sw_count})</div>
<table class="sw-table">
<thead><tr><th>#</th><th>Tên phần mềm</th><th>Phiên bản</th><th>Nhà phát hành</th><th>Ngày cài đặt</th></tr></thead>
<tbody>{sw_rows}</tbody>
</table>
</div>'''
            else:
                sw_table = '<div class="sw-section"><div class="sw-header">📦 Danh sách phần mềm (0)</div><div style="padding:8px 12px;color:#8892a4;font-size:11px;">(Chưa có dữ liệu)</div></div>'

            user_info_line = ""
            if user_info.get("user_name"):
                user_info_line = '<span style="color:#00d4aa;margin-left:12px;">👤 ' + _html_mod.escape(user_info.get("user_name", "")) + '</span>'
                if user_info.get("employee_id"):
                    user_info_line += ' <span style="color:#8892a4;margin-left:6px;">🪪 ' + _html_mod.escape(user_info.get("employee_id", "")) + '</span>'

            cards_html += f'''<div class="machine-card">
<div class="machine-header" onclick="var d=this.nextElementSibling;var a=this.querySelector('.arrow');if(d.style.display==='none'){{d.style.display='block';a.textContent='▼';}}else{{d.style.display='none';a.textContent='▶';}}">
<span class="dot" style="background:{status_colors.get(status, '#888')};"></span>
<strong>{_html_mod.escape(hostname)}</strong>{user_info_line}
<span style="margin-left:auto;color:{status_colors.get(status, '#888')};font-size:12px;font-weight:600;">{status}</span>
<span class="arrow" style="margin-left:8px;font-size:14px;color:#8892a4;">▶</span>
</div>
<div class="machine-detail" style="display:none;">
<table>
<tr><th>Machine ID</th><td>{mid}</td></tr>'''
            if user_info.get("user_name"):
                cards_html += f'<tr><th>Người dùng</th><td>{_html_mod.escape(user_info.get("user_name",""))} ({_html_mod.escape(user_info.get("employee_id",""))})</td></tr>'
            if user_info.get("email"):
                cards_html += f'<tr><th>Email</th><td>{_html_mod.escape(user_info.get("email",""))}</td></tr>'
            cards_html += f'''<tr><th>Hệ điều hành</th><td>{_html_mod.escape(f"{os_info.get('name','')} {os_info.get('release','')} (Build {os_info.get('build','')})".strip())}</td></tr>
<tr><th>CPU</th><td>{_html_mod.escape(cpu_info.get("name","") or "-")}</td></tr>
<tr><th>RAM</th><td>{ram_info.get("total_gb", 0) or 0} GB</td></tr>
<tr><th>Ổ đĩa</th><td>{total_disk if total_disk > 0 else "-"} GB</td></tr>
<tr><th>GPU</th><td>{_html_mod.escape(gpu_names) if gpu_names else "-"}</td></tr>
<tr><th>IP</th><td>{_html_mod.escape(m.get("ip_address","-"))}</td></tr>
<tr><th>Agent</th><td>{_html_mod.escape(m.get("version","-"))}</td></tr>
</table>
{sw_table}
</div>
</div>'''

        html = f'''<!DOCTYPE html>
<html lang="vi">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>GIAM-SAT - Báo cáo cấu hình máy trạm</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box;}}
body{{background:#0f1923;color:#eef4f8;font-family:'Segoe UI',sans-serif;font-size:13px;padding:20px;}}
.header{{background:#1a2a3a;border:1px solid #2a3a4a;border-radius:8px;padding:16px 20px;margin-bottom:16px;display:flex;justify-content:space-between;align-items:center;}}
.header h2{{color:#00d4aa;font-size:18px;}}
.header .meta{{color:#8892a4;font-size:12px;}}
.stats{{display:flex;gap:12px;margin-bottom:16px;}}
.stat{{background:#1a2a3a;border:1px solid #2a3a4a;border-radius:6px;padding:12px 20px;text-align:center;flex:1;}}
.stat .val{{font-size:24px;font-weight:700;color:#00d4aa;}}
.stat .lbl{{font-size:10px;color:#8892a4;text-transform:uppercase;letter-spacing:0.5px;}}
.machine-card{{background:#1a2a3a;border:1px solid #2a3a4a;border-radius:6px;margin-bottom:8px;overflow:hidden;}}
.machine-header{{display:flex;align-items:center;padding:12px 16px;cursor:pointer;user-select:none;transition:background 0.15s;}}
.machine-header:hover{{background:#2a3a4a;}}
.dot{{width:8px;height:8px;border-radius:50%;margin-right:10px;flex-shrink:0;}}
.machine-detail{{padding:0 16px 12px;}}
.machine-detail table{{width:100%;border-collapse:collapse;font-size:12px;}}
.machine-detail th{{text-align:left;padding:4px 8px;color:#6a8aaa;width:140px;font-weight:500;vertical-align:top;}}
.machine-detail td{{padding:4px 8px;color:#c8d8e8;vertical-align:top;}}
.sw-section{{margin-top:12px;border-top:1px solid #2a3a4a;}}
.sw-header{{padding:8px 12px;font-weight:600;font-size:12px;color:#00d4aa;background:rgba(0,212,170,0.05);}}
.sw-table{{width:100%;border-collapse:collapse;font-size:11px;}}
.sw-table th{{background:#1a3a3a;padding:6px 8px;text-align:left;color:#88cc99;font-weight:600;border-bottom:2px solid #2a5a3a;position:sticky;top:0;z-index:1;}}
.sw-table td{{padding:5px 8px;color:#c8d8e8;border-bottom:1px solid #1a2a3a;}}
.sw-table tbody tr:hover{{background:#2a3a4a;}}
.sw-table td:first-child{{color:#6a8aaa;width:30px;text-align:right;}}
.sw-table td:nth-child(2){{color:#eef4f8;font-weight:500;}}
.sw-table td:nth-child(5){{font-size:10px;color:#6a8aaa;}}
.footer{{margin-top:20px;text-align:center;color:#5a6a7a;font-size:10px;}}
@media print{{
body{{background:#fff;color:#000;}}
.machine-card{{border:1px solid #ccc;break-inside:avoid;}}
.machine-header{{background:#f0f0f0;color:#000;}}
.machine-detail{{display:block!important;}}
}}
</style>
</head>
<body>
<div class="header">
<div><h2>🖥 GIAM-SAT - Báo cáo cấu hình máy trạm</h2></div>
<div class="meta">📅 {now} | {len(machines)} máy ({online_count} Online)</div>
</div>
<div class="stats">
<div class="stat"><div class="val">{len(machines)}</div><div class="lbl">Tổng máy</div></div>
<div class="stat"><div class="val">{online_count}</div><div class="lbl">Online</div></div>
<div class="stat"><div class="val">{len(machines) - online_count}</div><div class="lbl">Offline</div></div>
<div class="stat"><div class="val">{now}</div><div class="lbl">Ngày xuất</div></div>
</div>
{cards_html}
<div class="footer">🔒 GIAM-SAT v2.5.1 - Xuất báo cáo lúc {now}</div>
</body>
</html>'''

        core.db.insert_audit_log(username, "export_config_html",
            f"Exported {len(machines)} machines config to HTML", request.remote_addr)

        return Response(html, mimetype="text/html; charset=utf-8",
                      headers={"Content-Disposition": f"attachment; filename=GIAM-SAT_Config_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"})

    # Route moved to api_cleanup.py (v3.2) - handles selective cleanup by type
    # @app.route("/api/cleanup", methods=["POST"])
    # def api_cleanup():
    #     username, err, code = check_auth("settings")
    #     if err: return err, code
    #     data = request.json or {}
    #     days = data.get("days", 30)
    #     keep_threats = data.get("keep_threats", True)
    #     deleted = core.db.cleanup_old_logs(days=days, keep_threats=keep_threats)
    #     core.db.insert_audit_log(username, "cleanup_logs", f"Deleted {deleted} records older than {days} days", request.remote_addr)
    #     print(f"[*] Cleanup: deleted {deleted} old records by {username}")
    #     return jsonify({"success": True, "deleted": deleted})

    @app.route("/api/alerting/config", methods=["GET"])
    def api_alerting_config():
        username, err, code = check_auth("settings")
        if err: return err, code
        return jsonify(core.alerting.config)

    @app.route("/api/alerting/config", methods=["POST"])
    def api_alerting_update_config():
        username, err, code = check_auth("settings")
        if err: return err, code
        data = request.json
        for key, value in data.items():
            core.alerting.set_config(key, value)
        core.db.insert_audit_log(username, "update_alerting", json.dumps(data), request.remote_addr)
        return jsonify({"success": True})