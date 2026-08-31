"""
API Assets - v4.4: Quản lý tài sản (máy tính, màn hình)
REST endpoints cho danh sách tài sản và lịch sử thay đổi.
"""
import json
from flask import Blueprint, request, jsonify
from .api_common import check_auth

assets_bp = Blueprint("api_assets", __name__)


def init_assets_api(app, db):
    """Register asset management routes on the Flask app."""

    @app.route("/api/assets/computers")
    def api_assets_computers():
        _, err, code = check_auth("api")
        if err: return err, code
        search = request.args.get("search", "").strip()
        limit = int(request.args.get("limit", 200))
        rows = db.get_asset_computers(search=search, limit=limit) if db else []
        return jsonify({"computers": rows})

    @app.route("/api/assets/monitors")
    def api_assets_monitors():
        _, err, code = check_auth("api")
        if err: return err, code
        search = request.args.get("search", "").strip()
        limit = int(request.args.get("limit", 200))
        rows = db.get_asset_monitors(search=search, limit=limit) if db else []
        return jsonify({"monitors": rows})

    # =========================================================================
    # v4.7: IT asset inventory (Kho - manual + auto-discovered)
    # =========================================================================
    def _inv_method(name):
        """Return bound db method if available, else None (safe for all backends)."""
        return getattr(db, name, None) if db else None

    @app.route("/api/assets/inventory")
    def api_assets_inventory():
        _, err, code = check_auth("api")
        if err: return err, code
        m = _inv_method("get_asset_inventory")
        if not m:
            return jsonify({"assets": []})
        category = request.args.get("category", "").strip() or None
        status = request.args.get("status", "").strip() or None
        source = request.args.get("source", "").strip() or None
        search = request.args.get("search", "").strip() or None
        # v5.0.4 (MEDIUM-19): never let invalid limit values raise a 500
        try:
            limit = max(1, min(int(request.args.get("limit", 500)), 5000))
        except (TypeError, ValueError):
            limit = 500
        rows = m(category=category, status=status, source=source, search=search, limit=limit)
        return jsonify({"assets": rows})

    @app.route("/api/assets/inventory/stats")
    def api_assets_inventory_stats():
        _, err, code = check_auth("api")
        if err: return err, code
        m = _inv_method("get_asset_inventory_stats")
        if not m:
            return jsonify({"by_category": [], "by_status": [], "total": 0})
        return jsonify(m())

    @app.route("/api/assets/inventory", methods=["POST"])
    def api_assets_inventory_add():
        username, err, code = check_auth("delete")
        if err: return err, code
        m = _inv_method("upsert_inventory_asset")
        if not m:
            return jsonify({"error": "DB method unavailable"}), 500
        data = request.json or {}
        if not data.get("category"):
            return jsonify({"error": "Thiếu category"}), 400
        data["source"] = data.get("source") or "manual"
        result = m(data)
        db.insert_audit_log(username, "asset_add",
            f"Thêm tài sản category='{data.get('category')}' name='{data.get('name') or ''}'",
            request.remote_addr)
        return jsonify({"success": bool(result), **result}), 201 if result else 500

    @app.route("/api/assets/inventory/<asset_id>", methods=["PUT"])
    def api_assets_inventory_update(asset_id):
        username, err, code = check_auth("delete")
        if err: return err, code
        m = _inv_method("upsert_inventory_asset")
        if not m:
            return jsonify({"error": "DB method unavailable"}), 500
        data = request.json or {}
        data["asset_id"] = asset_id
        data["source"] = data.get("source") or "manual"
        result = m(data)
        db.insert_audit_log(username, "asset_update",
            f"Cập nhật tài sản asset_id='{asset_id}'", request.remote_addr)
        return jsonify({"success": bool(result), **result})

    @app.route("/api/assets/inventory/<asset_id>", methods=["DELETE"])
    def api_assets_inventory_delete(asset_id):
        username, err, code = check_auth("delete")
        if err: return err, code
        m = _inv_method("delete_inventory_asset")
        if not m:
            return jsonify({"error": "DB method unavailable"}), 500
        ok = m(asset_id)
        if ok:
            db.insert_audit_log(username, "asset_delete",
                f"Xóa tài sản asset_id='{asset_id}'", request.remote_addr)
        return jsonify({"success": ok})

    @app.route("/api/assets/inventory/<asset_id>/adopt", methods=["POST"])
    def api_assets_inventory_adopt(asset_id):
        """Chuyển tài sản auto->manual: gán owner/location/mã TS."""
        username, err, code = check_auth("delete")
        if err: return err, code
        m = _inv_method("adopt_inventory_asset")
        if not m:
            return jsonify({"error": "DB method unavailable"}), 500
        data = request.json or {}
        ok = m(asset_id, data)
        if ok:
            db.insert_audit_log(username, "asset_adopt",
                f"Adopt tài sản asset_id='{asset_id}'", request.remote_addr)
        return jsonify({"success": ok})

    @app.route("/api/assets/users/sync", methods=["POST"])
    def api_assets_users_sync():
        """Đồng bộ danh sách người dùng (họ tên, mã NV, email) từ assets_computers -> assets_inventory category=user."""
        _, err, code = check_auth("delete")
        if err: return err, code
        m = _inv_method("sync_user_assets")
        if not m:
            return jsonify({"error": "DB method unavailable"}), 500
        try:
            created = m()
            return jsonify({"success": True, "created": created})
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    @app.route("/api/assets/discovery/scan", methods=["POST"])
    def api_assets_discovery_scan():
        """Quét dải IP bằng SNMP + port fingerprint -> tự nạp tài sản. (Admin only)"""
        _, err, code = check_auth("delete")
        if err: return err, code
        data = request.json or {}
        cidr = (data.get("range") or "").strip()
        if not cidr:
            return jsonify({"error": "Thiếu dải IP (vd 192.168.1.0/24)"}), 400
        try:
            from asset_discovery import parse_cidr, run_scan
        except Exception:
            return jsonify({"error": "Module asset_discovery không tải được"}), 500
        ips = parse_cidr(cidr)
        if len(ips) > 4096:
            return jsonify({"error": "Dải IP quá lớn (tối đa 4096 địa chỉ, VD /22). Khuyến nghị quét /24."}), 400
        try:
            summary = run_scan(cidr, db, max_threads=32, timeout=0.4)
            return jsonify(summary)
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    @app.route("/api/assets/changes")
    def api_assets_changes():
        _, err, code = check_auth("api")
        if err: return err, code
        limit = int(request.args.get("limit", 100))
        unresolved_only = request.args.get("unresolved", "0") == "1"
        rows = db.get_asset_change_log(limit=limit, unresolved_only=unresolved_only) if db else []
        return jsonify({"changes": rows})

    @app.route("/api/assets/changes/<int:change_id>/resolve", methods=["POST"])
    def api_assets_resolve_change(change_id):
        # v4.10 (LOW-5): resolving change is an admin action; resolved_by comes
        # from the authenticated token, not a client-supplied field.
        username, err, code = check_auth("settings")
        if err: return err, code
        resolved_by = username or "admin"
        ok = db.resolve_asset_change(change_id, resolved_by) if db else False
        return jsonify({"success": ok})

    @app.route("/api/assets/unresolved_count")
    def api_assets_unresolved_count():
        """Return count of unresolved asset changes for badge display."""
        _, err, code = check_auth("api")
        if err: return err, code
        m = _inv_method("get_asset_change_log")
        if not m:
            return jsonify({"count": 0})
        try:
            rows = m(limit=100000, unresolved_only=True) or []
            return jsonify({"count": len(rows)})
        except Exception:
            return jsonify({"count": 0})

    @app.route("/api/assets/export")
    def api_assets_export():
        """Export all assets as Excel file with 2 sheets."""
        _, err, code = check_auth("api")
        if err: return err, code
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from io import BytesIO
        except ImportError:
            return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500

        wb = openpyxl.Workbook()
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1a3a5a", end_color="1a3a5a", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Sheet 1: May tinh
        ws1 = wb.active
        ws1.title = "May tinh"
        pc_headers = ["Ma TS", "May tinh", "Nguoi dung", "MNSV", "Email",
                       "Mainboard", "Mainboard Serial", "CPU", "Cores", "Clock MHz",
                       "RAM (GB)", "O cung", "GPU", "Man hinh", "OS", "Online", "Cap nhat"]
        for col, h in enumerate(pc_headers, 1):
            cell = ws1.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        computers = db.get_asset_computers(limit=5000) if db else []
        for r, c in enumerate(computers, 2):
            row_data = [
                c.get('display_id') or c.get('asset_id', '-'),
                c.get('hostname') or c.get('machine_id', '-'),
                c.get('user_name', '-'),
                c.get('employee_id', '-'),
                c.get('email', '-'),
                (c.get('motherboard_manufacturer','') + ' ' + c.get('motherboard_product','')).strip() or '-',
                c.get('motherboard_serial', '-'),
                c.get('cpu_name', '-'),
                c.get('cpu_cores', 0),
                c.get('cpu_max_clock_mhz', 0),
                c.get('ram_total_gb', 0),
                '',
                '',
                '',
                (c.get('os_name','') + ' ' + c.get('os_version','')).strip() or '-',
                'Online' if c.get('is_online') else 'Offline',
                str(c.get('updated_at', ''))[:19],
            ]
            # Parse disks
            disks = c.get('disks_json', [])
            if isinstance(disks, str):
                try: disks = json.loads(disks)
                except: disks = []
            row_data[11] = '; '.join([f"{d.get('model','?')} ({d.get('size_gb','?')}GB {d.get('interface','')})" for d in disks]) or '-'
            # Parse GPU
            gpus = c.get('gpu_json', [])
            if isinstance(gpus, str):
                try: gpus = json.loads(gpus)
                except: gpus = []
            row_data[12] = '; '.join([f"{g.get('name','?')} ({g.get('ram_gb','?')}GB)" for g in gpus]) or '-'
            # Parse monitors
            monitors_json = c.get('monitors_json', [])
            if isinstance(monitors_json, str):
                try: monitors_json = json.loads(monitors_json)
                except: monitors_json = []
            row_data[13] = '; '.join([f"{m.get('manufacturer','')} {m.get('name','')} ({m.get('resolution','')})" for m in monitors_json]) or '-'
            for col, val in enumerate(row_data, 1):
                cell = ws1.cell(row=r, column=col, value=val)
                cell.border = thin_border

        # Sheet 2: Man hinh
        ws2 = wb.create_sheet("Man hinh")
        mn_headers = ["Ma TS", "Ten man hinh", "Hang", "Do phan giai", "Loai", "May tinh ket noi", "Nguoi dung", "Cap nhat"]
        for col, h in enumerate(mn_headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        monitors = db.get_asset_monitors(limit=5000) if db else []
        for r, m in enumerate(monitors, 2):
            row_data = [
                m.get('display_id') or m.get('asset_id', '-'),
                m.get('name', '-'),
                m.get('manufacturer', '-'),
                m.get('resolution', '-'),
                m.get('model_type', 'Monitor'),
                m.get('computer_hostname', 'Chua gan'),
                m.get('computer_user', '-'),
                str(m.get('updated_at', ''))[:19],
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws2.cell(row=r, column=col, value=val)
                cell.border = thin_border

        # =====================================================================
        # v4.7: IT asset inventory sheets (Kho + auto-discovered devices)
        # =====================================================================
        inv_headers = ["Mã TS", "Loại", "Tên", "Hãng", "Model", "Serial",
                       "Status", "Gán cho", "IP", "Vị trí", "Mua", "Bảo hành",
                       "Giá", "Nguồn", "Ghi chú", "Cập nhật"]
        inv_rows = db.get_asset_inventory(limit=5000) if (db and hasattr(db, 'get_asset_inventory')) else []
        status_map = {"in_stock": "Còn hàng", "online": "Online", "assigned": "Đã cấp",
                      "in_repair": "Đang sửa", "disposed": "Thanh lý"}

        def _fill_inventory_sheet(sheet, items):
            for col, h in enumerate(inv_headers, 1):
                cell = sheet.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            for r, a in enumerate(items, 2):
                source_label = 'Tự động' if a.get('source') == 'auto' else 'Nhập tay'
                row_data = [
                    a.get('display_id') or a.get('asset_id', '-'),
                    a.get('category', '-'),
                    a.get('name', '-'),
                    a.get('brand', '-'),
                    a.get('model', '-'),
                    a.get('serial_number', '-'),
                    status_map.get(a.get('status'), a.get('status', '-')),
                    a.get('assigned_to', '-'),
                    a.get('ip_address', '-'),
                    a.get('location', '-'),
                    a.get('purchase_date', '-'),
                    a.get('warranty_until', '-'),
                    a.get('cost', 0),
                    a.get('quantity', 1),
                    source_label,
                    a.get('notes', '-'),
                    str(a.get('updated_at', ''))[:19],
                ]
                for col, val in enumerate(row_data, 1):
                    sheet.cell(row=r, column=col, value=val).border = thin_border

        def _inv_items(cat=None, manual_only=False):
            if manual_only:
                return [a for a in inv_rows if a.get("source") == "manual"]
            return [a for a in inv_rows if a.get("category") == cat]

        _fill_inventory_sheet(wb.create_sheet("May in"), _inv_items("printer"))
        _fill_inventory_sheet(wb.create_sheet("Dien thoai"), _inv_items("phone"))
        _fill_inventory_sheet(wb.create_sheet("Thiet bi mang"), _inv_items("network_device"))
        _fill_inventory_sheet(wb.create_sheet("Ngoai vi"), _inv_items("peripheral"))
        _fill_inventory_sheet(wb.create_sheet("Kho"), _inv_items(manual_only=True))

        # v4.9: Nguoi dung sheet (digital user assets)
        user_items = [a for a in inv_rows if a.get("category") == "user"]
        ws_users = wb.create_sheet("Nguoi dung")
        uheaders = ["Ho ten", "Ma NV", "Email", "Tai khoan noi bo", "Van phong/chi nhanh", "Nguon", "Cap nhat"]
        for col, h in enumerate(uheaders, 1):
            cell = ws_users.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        for r, a in enumerate(user_items, 2):
            email = a.get('email', '')
            internal = email.split('@')[0] if email else ''
            source_label = 'Tu dong' if a.get('source') == 'auto' else 'Nhap tay'
            udata = [a.get('name', ''), a.get('employee_id', ''), email, internal,
                     a.get('location', ''), source_label, str(a.get('updated_at', ''))[:19]]
            for col, val in enumerate(udata, 1):
                ws_users.cell(row=r, column=col, value=val).border = thin_border

        # Auto-width
        for ws in wb.worksheets:
            for col in ws.columns:
                max_len = 0
                for cell in col:
                    try:
                        max_len = max(max_len, len(str(cell.value or '')))
                    except:
                        pass
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        from flask import send_file
        timestamp = __import__('datetime').datetime.now().strftime("%Y%m%d_%H%M%S")
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'GIAM-SAT_TaiSan_{timestamp}.xlsx'
        )

    print("[*] API Assets registered: /api/assets/*")
